"""
Generator Excel: Tanda Tangan Digital & Verifikasi (RSA, El Gamal, Schnorr)
Semua bilangan kunci berada di rentang 2000-3000 sesuai permintaan tugas.
Proses pangkat-modulo (modpow) ditampilkan sebagai tabel Square-and-Multiply
dengan RUMUS EXCEL ASLI (bukan hasil hardcode dari Python), karena modulus
< 3000 sehingga hasil kuadrat (maks ~9 juta) masih aman dari overflow Excel.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FONT = Font(bold=True)
TITLE_FILL = PatternFill("solid", start_color="1F4E78")
SECTION_FILL = PatternFill("solid", start_color="2E75B6")
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
RESULT_FILL = PatternFill("solid", start_color="FFF2CC")
thin = Side(style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(ws, ref, font=None, fill=None, align=None, border=False, fmt=None):
    c = ws[ref]
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = Alignment(horizontal=align)
    if border: c.border = BORDER
    if fmt: c.number_format = fmt
    return c

def title_bar(ws, row, text, span=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 2

def section_bar(ws, row, text, span=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    return row + 1

def kv(ws, row, label, value, is_formula=False, note=None, fmt=None):
    ws.cell(row=row, column=1, value=label).font = BOLD
    c = ws.cell(row=row, column=2, value=value)
    c.font = BLACK if is_formula else BLUE
    c.border = BORDER
    if fmt: c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, color="808080", size=9)
    return row + 1

def modpow_table(ws, start_row, title, base_ref, exp_ref, mod_ref, bitlen, base_val_hint=None):
    """
    Writes a Square-and-Multiply modular exponentiation table using live Excel
    formulas. base_ref/exp_ref/mod_ref are cell references (e.g. '$B$4').
    Returns (result_cell_ref, next_free_row).
    """
    r = start_row
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    headers = ["Bit ke-i (MSB→LSB)", "2^i", "Bit eksponen (b_i)", "Rumus Hasil Sementara (Square-and-Multiply)", "Hasil"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    r += 1

    # initial row: result = 1 (before processing MSB)
    init_row = r
    ws.cell(row=r, column=1, value="awal").border = BORDER
    ws.cell(row=r, column=2, value="-").border = BORDER
    ws.cell(row=r, column=3, value="-").border = BORDER
    ws.cell(row=r, column=4, value="Hasil awal = 1").border = BORDER
    res_cell = ws.cell(row=r, column=5, value=1)
    res_cell.border = BORDER
    res_cell.font = BLACK
    prev_result_row = r
    r += 1

    first_data_row = r
    for idx, i in enumerate(range(bitlen - 1, -1, -1)):
        bit_cell_ref = f"C{r}"
        prev_res_ref = f"E{prev_result_row}"
        ws.cell(row=r, column=1, value=i).border = BORDER
        ws.cell(row=r, column=2, value=f"=2^{i}").border = BORDER
        ws.cell(row=r, column=2).font = BLACK
        bit_formula = f"=MOD(INT({exp_ref}/B{r}),2)"
        bc = ws.cell(row=r, column=3, value=bit_formula)
        bc.font = BLACK
        bc.border = BORDER
        # square then conditional multiply, all in one live formula
        formula = (
            f'=IF(C{r}=1,'
            f'MOD(MOD({prev_res_ref}^2,{mod_ref})*{base_ref},{mod_ref}),'
            f'MOD({prev_res_ref}^2,{mod_ref}))'
        )
        desc = ws.cell(row=r, column=4,
                        value=f"IF(bit=1, (hasil_prev^2 mod n)*basis mod n, hasil_prev^2 mod n)")
        desc.border = BORDER
        desc.font = Font(italic=True, size=9, color="555555")
        rc = ws.cell(row=r, column=5, value=formula)
        rc.font = BLACK
        rc.border = BORDER
        prev_result_row = r
        r += 1

    final_ref = f"E{prev_result_row}"
    return final_ref, r + 1

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

wb = Workbook()

# =========================================================================
# SHEET 1: RSA
# =========================================================================
ws = wb.active
ws.title = "RSA"
set_widths(ws, {"A": 26, "B": 16, "C": 50, "D": 14, "E": 14})

r = title_bar(ws, 1, "RSA — Tanda Tangan Digital dan Verifikasi")
r = section_bar(ws, r, "1. Pembangkitan Kunci")
r = kv(ws, r, "p (bilangan prima)", 47, note="input")
r = kv(ws, r, "q (bilangan prima)", 59, note="input")
n_row = r
r = kv(ws, r, "n = p × q", "=B4*B5", is_formula=True, note="2000 ≤ n ≤ 3000")
phi_row = r
r = kv(ws, r, "φ(n) = (p-1)(q-1)", "=(B4-1)*(B5-1)", is_formula=True)
e_row = r
r = kv(ws, r, "e (kunci publik)", 17, note="dipilih, gcd(e,φ(n))=1")
r = kv(ws, r, "Cek gcd(e, φ(n))", "=GCD(B8,B7)", is_formula=True, note="harus = 1")
d_row = r
r = kv(ws, r, "d (kunci privat)", 157, note="d = e^-1 mod φ(n), via Extended Euclidean Algorithm")
r = kv(ws, r, "Verifikasi d: (e × d) mod φ(n)", "=MOD(B8*B10,B7)", is_formula=True, note="harus = 1")
r += 1
ws.cell(row=r, column=1, value="Kunci Publik  : (n, e) =").font = BOLD
ws.cell(row=r, column=2, value="=\"(\"&B6&\", \"&B8&\")\"").font = BLACK
r += 1
ws.cell(row=r, column=1, value="Kunci Privat  : (n, d) =").font = BOLD
ws.cell(row=r, column=2, value="=\"(\"&B6&\", \"&B10&\")\"").font = BLACK
r += 2

r = section_bar(ws, r, "2. Pesan yang Akan Ditandatangani")
m_row = r
r = kv(ws, r, "M (pesan, M < n)", 1234, note="input")
r += 1

r = section_bar(ws, r, "3. Proses Penandatanganan: S = M^d mod n")
sig_final, r = modpow_table(ws, r, "Square-and-Multiply  (basis = M, eksponen = d, modulus = n)",
                             base_ref=f"$B${m_row}", exp_ref=f"$B${d_row}", mod_ref=f"$B${n_row}", bitlen=8)
S_row = r
ws.cell(row=r, column=1, value="Tanda Tangan Digital  S =").font = BOLD
sc = ws.cell(row=r, column=2, value=f"={sig_final}")
sc.font = BLACK
sc.fill = RESULT_FILL
sc.border = BORDER
r += 2

r = section_bar(ws, r, "4. Proses Verifikasi: M' = S^e mod n")
ver_final, r = modpow_table(ws, r, "Square-and-Multiply  (basis = S, eksponen = e, modulus = n)",
                             base_ref=f"$B${S_row}", exp_ref=f"$B${e_row}", mod_ref=f"$B${n_row}", bitlen=5)
Mp_row = r
ws.cell(row=r, column=1, value="Pesan Hasil Verifikasi  M' =").font = BOLD
mc = ws.cell(row=r, column=2, value=f"={ver_final}")
mc.font = BLACK
mc.fill = RESULT_FILL
mc.border = BORDER
r += 1
ws.cell(row=r, column=1, value="Status Verifikasi (M' = M ?)").font = BOLD
status = ws.cell(row=r, column=2, value=f'=IF(B{Mp_row}=B{m_row},"VALID","TIDAK VALID")')
status.font = Font(bold=True, color="008000")
status.border = BORDER

# =========================================================================
# SHEET 2: EL GAMAL
# =========================================================================
ws2 = wb.create_sheet("El Gamal")
set_widths(ws2, {"A": 30, "B": 16, "C": 55, "D": 14, "E": 14})

r = title_bar(ws2, 1, "El Gamal — Tanda Tangan Digital dan Verifikasi")
r = section_bar(ws2, r, "1. Pembangkitan Kunci")
p_row = r
r = kv(ws2, r, "p (bilangan prima)", 2357, note="2000 ≤ p ≤ 3000")
g_row = r
r = kv(ws2, r, "g (akar primitif mod p)", 2, note="input")
x_row = r
r = kv(ws2, r, "x (kunci privat)", 15, note="1 < x < p-1, input")
pm1_row = r
r = kv(ws2, r, "p - 1", f"=B{p_row}-1", is_formula=True)
r += 1

r = section_bar(ws2, r, "Hitung y = g^x mod p (kunci publik)")
y_final, r = modpow_table(ws2, r, "Square-and-Multiply  (basis = g, eksponen = x, modulus = p)",
                           base_ref=f"$B${g_row}", exp_ref=f"$B${x_row}", mod_ref=f"$B${p_row}", bitlen=4)
y_row = r
ws2.cell(row=r, column=1, value="Kunci Publik  y = g^x mod p =").font = BOLD
yc = ws2.cell(row=r, column=2, value=f"={y_final}")
yc.font = BLACK; yc.fill = RESULT_FILL; yc.border = BORDER
r += 2

r = section_bar(ws2, r, "2. Pesan yang Akan Ditandatangani")
m_row = r
r = kv(ws2, r, "m (pesan, m < p-1)", 1000, note="input")
k_row = r
r = kv(ws2, r, "k (bilangan acak rahasia)", 13, note="gcd(k, p-1) = 1, input")
r = kv(ws2, r, "Cek gcd(k, p-1)", f"=GCD(B{k_row},B{pm1_row})", is_formula=True, note="harus = 1")
kinv_row = r
r = kv(ws2, r, "k^-1 mod (p-1)", 725, note="invers modulo k, via Extended Euclidean Algorithm")
r = kv(ws2, r, "Verifikasi: (k × k^-1) mod (p-1)", f"=MOD(B{k_row}*B{kinv_row},B{pm1_row})", is_formula=True, note="harus = 1")
r += 1

r = section_bar(ws2, r, "3. Proses Penandatanganan: r = g^k mod p ;  s = (m - x·r)·k^-1 mod (p-1)")
r_final, r = modpow_table(ws2, r, "Square-and-Multiply  (basis = g, eksponen = k, modulus = p)",
                           base_ref=f"$B${g_row}", exp_ref=f"$B${k_row}", mod_ref=f"$B${p_row}", bitlen=4)
r_row = r
ws2.cell(row=r, column=1, value="r = g^k mod p =").font = BOLD
rc = ws2.cell(row=r, column=2, value=f"={r_final}")
rc.font = BLACK; rc.fill = RESULT_FILL; rc.border = BORDER
r += 1
s_row = r
ws2.cell(row=r, column=1, value="s = (m - x·r) × k^-1 mod (p-1) =").font = BOLD
sc2 = ws2.cell(row=r, column=2, value=f"=MOD((B{m_row}-B{x_row}*B{r_row})*B{kinv_row},B{pm1_row})")
sc2.font = BLACK; sc2.fill = RESULT_FILL; sc2.border = BORDER
r += 1
ws2.cell(row=r, column=1, value="Tanda Tangan Digital = (r, s) =").font = BOLD
ws2.cell(row=r, column=2, value=f'=\"(\"&B{r_row}&\", \"&B{s_row}&\")\"').font = BLACK
r += 2

r = section_bar(ws2, r, "4. Proses Verifikasi: cek  g^m mod p  =  (y^r × r^s) mod p")
lhs_final, r = modpow_table(ws2, r, "Ruas kiri: Square-and-Multiply  (basis = g, eksponen = m, modulus = p)",
                             base_ref=f"$B${g_row}", exp_ref=f"$B${m_row}", mod_ref=f"$B${p_row}", bitlen=10)
lhs_row = r
ws2.cell(row=r, column=1, value="Ruas Kiri = g^m mod p =").font = BOLD
lc = ws2.cell(row=r, column=2, value=f"={lhs_final}")
lc.font = BLACK; lc.fill = RESULT_FILL; lc.border = BORDER
r += 2

yr_final, r = modpow_table(ws2, r, "Ruas kanan bagian-1: Square-and-Multiply  (basis = y, eksponen = r, modulus = p)",
                            base_ref=f"$B${y_row}", exp_ref=f"$B${r_row}", mod_ref=f"$B${p_row}", bitlen=11)
yr_row = r
ws2.cell(row=r, column=1, value="y^r mod p =").font = BOLD
yrc = ws2.cell(row=r, column=2, value=f"={yr_final}")
yrc.font = BLACK; yrc.border = BORDER
r += 2

rs_final, r = modpow_table(ws2, r, "Ruas kanan bagian-2: Square-and-Multiply  (basis = r, eksponen = s, modulus = p)",
                            base_ref=f"$B${r_row}", exp_ref=f"$B${s_row}", mod_ref=f"$B${p_row}", bitlen=10)
rs_row = r
ws2.cell(row=r, column=1, value="r^s mod p =").font = BOLD
rsc = ws2.cell(row=r, column=2, value=f"={rs_final}")
rsc.font = BLACK; rsc.border = BORDER
r += 1
rhs_row = r
ws2.cell(row=r, column=1, value="Ruas Kanan = (y^r × r^s) mod p =").font = BOLD
rhc = ws2.cell(row=r, column=2, value=f"=MOD(B{yr_row}*B{rs_row},B{p_row})")
rhc.font = BLACK; rhc.fill = RESULT_FILL; rhc.border = BORDER
r += 1
ws2.cell(row=r, column=1, value="Status Verifikasi (Kiri = Kanan ?)").font = BOLD
status2 = ws2.cell(row=r, column=2, value=f'=IF(B{lhs_row}=B{rhs_row},"VALID","TIDAK VALID")')
status2.font = Font(bold=True, color="008000")
status2.border = BORDER

# =========================================================================
# SHEET 3: SCHNORR
# =========================================================================
ws3 = wb.create_sheet("Schnorr")
set_widths(ws3, {"A": 32, "B": 16, "C": 55, "D": 14, "E": 14})

r = title_bar(ws3, 1, "Schnorr — Tanda Tangan Digital dan Verifikasi")
r = section_bar(ws3, r, "1. Pembangkitan Kunci")
p_row = r
r = kv(ws3, r, "p (bilangan prima)", 2357, note="2000 ≤ p ≤ 3000")
q_row = r
r = kv(ws3, r, "q (prima pembagi p-1)", 31, note="p-1 = 2356 = 2^2 × 19 × 31")
h_row = r
r = kv(ws3, r, "h (basis bantu, 1 < h < p)", 3, note="input")
pm1_row = r
r = kv(ws3, r, "(p-1)/q (eksponen pembentuk g)", 76, note="=2356/31")
r += 1

r = section_bar(ws3, r, "Hitung g = h^((p-1)/q) mod p  (generator orde q)")
g_final, r = modpow_table(ws3, r, "Square-and-Multiply  (basis = h, eksponen = (p-1)/q, modulus = p)",
                           base_ref=f"$B${h_row}", exp_ref=f"$B${pm1_row}", mod_ref=f"$B${p_row}", bitlen=7)
g_row = r
ws3.cell(row=r, column=1, value="g = h^((p-1)/q) mod p =").font = BOLD
gc = ws3.cell(row=r, column=2, value=f"={g_final}")
gc.font = BLACK; gc.fill = RESULT_FILL; gc.border = BORDER
r += 1
x_row = r
r = kv(ws3, r, "x (kunci privat)", 7, note="1 < x < q, input")
r += 1

r = section_bar(ws3, r, "Hitung y = g^x mod p  (kunci publik)")
y_final, r = modpow_table(ws3, r, "Square-and-Multiply  (basis = g, eksponen = x, modulus = p)",
                           base_ref=f"$B${g_row}", exp_ref=f"$B${x_row}", mod_ref=f"$B${p_row}", bitlen=3)
y_row = r
ws3.cell(row=r, column=1, value="Kunci Publik  y = g^x mod p =").font = BOLD
yc = ws3.cell(row=r, column=2, value=f"={y_final}")
yc.font = BLACK; yc.fill = RESULT_FILL; yc.border = BORDER
r += 2

r = section_bar(ws3, r, "2. Pesan yang Akan Ditandatangani")
m_row = r
r = kv(ws3, r, "m (pesan)", 19, note="input")
k_row = r
r = kv(ws3, r, "k (bilangan acak rahasia, 1<k<q)", 5, note="input")
r += 1

r = section_bar(ws3, r, "3. Proses Penandatanganan: r = g^k mod p ; e = (r+m) mod q ; s = (k+x·e) mod q")
r_final, r = modpow_table(ws3, r, "Square-and-Multiply  (basis = g, eksponen = k, modulus = p)",
                           base_ref=f"$B${g_row}", exp_ref=f"$B${k_row}", mod_ref=f"$B${p_row}", bitlen=3)
r_row = r
ws3.cell(row=r, column=1, value="r = g^k mod p =").font = BOLD
rc = ws3.cell(row=r, column=2, value=f"={r_final}")
rc.font = BLACK; rc.fill = RESULT_FILL; rc.border = BORDER
r += 1
e_row = r
ws3.cell(row=r, column=1, value="e = (r + m) mod q   [fungsi hash sederhana]").font = BOLD
ec = ws3.cell(row=r, column=2, value=f"=MOD(B{r_row}+B{m_row},B{q_row})")
ec.font = BLACK; ec.fill = RESULT_FILL; ec.border = BORDER
r += 1
s_row = r
ws3.cell(row=r, column=1, value="s = (k + x × e) mod q =").font = BOLD
sc3 = ws3.cell(row=r, column=2, value=f"=MOD(B{k_row}+B{x_row}*B{e_row},B{q_row})")
sc3.font = BLACK; sc3.fill = RESULT_FILL; sc3.border = BORDER
r += 1
ws3.cell(row=r, column=1, value="Tanda Tangan Digital = (s, e) =").font = BOLD
ws3.cell(row=r, column=2, value=f'=\"(\"&B{s_row}&\", \"&B{e_row}&\")\"').font = BLACK
r += 2

r = section_bar(ws3, r, "4. Proses Verifikasi: r' = g^s × y^(q-e) mod p ; e' = (r'+m) mod q ; cek e' = e")
gs_final, r = modpow_table(ws3, r, "Bagian-1: Square-and-Multiply  (basis = g, eksponen = s, modulus = p)",
                            base_ref=f"$B${g_row}", exp_ref=f"$B${s_row}", mod_ref=f"$B${p_row}", bitlen=3)
gs_row = r
ws3.cell(row=r, column=1, value="g^s mod p =").font = BOLD
gsc = ws3.cell(row=r, column=2, value=f"={gs_final}")
gsc.font = BLACK; gsc.border = BORDER
r += 1
qe_row = r
ws3.cell(row=r, column=1, value="(q - e)  [eksponen invers y]").font = BOLD
qec = ws3.cell(row=r, column=2, value=f"=B{q_row}-B{e_row}")
qec.font = BLACK; qec.border = BORDER
r += 2

yqe_final, r = modpow_table(ws3, r, "Bagian-2: Square-and-Multiply  (basis = y, eksponen = (q-e), modulus = p)",
                             base_ref=f"$B${y_row}", exp_ref=f"$B${qe_row}", mod_ref=f"$B${p_row}", bitlen=4)
yqe_row = r
ws3.cell(row=r, column=1, value="y^(q-e) mod p =").font = BOLD
yqec = ws3.cell(row=r, column=2, value=f"={yqe_final}")
yqec.font = BLACK; yqec.border = BORDER
r += 1
rv_row = r
ws3.cell(row=r, column=1, value="r' = (g^s × y^(q-e)) mod p =").font = BOLD
rvc = ws3.cell(row=r, column=2, value=f"=MOD(B{gs_row}*B{yqe_row},B{p_row})")
rvc.font = BLACK; rvc.fill = RESULT_FILL; rvc.border = BORDER
r += 1
ev_row = r
ws3.cell(row=r, column=1, value="e' = (r' + m) mod q =").font = BOLD
evc = ws3.cell(row=r, column=2, value=f"=MOD(B{rv_row}+B{m_row},B{q_row})")
evc.font = BLACK; evc.fill = RESULT_FILL; evc.border = BORDER
r += 1
ws3.cell(row=r, column=1, value="Status Verifikasi (e' = e ?)").font = BOLD
status3 = ws3.cell(row=r, column=2, value=f'=IF(B{ev_row}=B{e_row},"VALID","TIDAK VALID")')
status3.font = Font(bold=True, color="008000")
status3.border = BORDER

# global font
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.font.name != "Arial":
                cell.font = Font(name="Arial", bold=cell.font.bold, italic=cell.font.italic,
                                  color=cell.font.color, size=cell.font.size or 10)

wb.save("Tanda_Tangan_Digital_RSA_ElGamal_Schnorr.xlsx")
print("saved")